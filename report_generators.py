#!/usr/bin/env python3
"""
Fixed Excel report generator with robust error handling and data type safety.
"""

import io
import logging
from typing import List, Dict, Any
from openpyxl.styles import Font, PatternFill

def safe_float(value, default=0.0):
    """Safely convert value to float."""
    try:
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            # Remove currency symbols and whitespace
            clean_value = value.replace('₹', '').replace(',', '').strip()
            if clean_value:
                return float(clean_value)
        return default
    except (ValueError, TypeError):
        return default

def safe_int(value, default=0):
    """Safely convert value to int."""
    try:
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return int(value)
        if isinstance(value, str):
            clean_value = value.replace(',', '').strip()
            if clean_value:
                return int(float(clean_value))
        return default
    except (ValueError, TypeError):
        return default

def safe_str(value, default=""):
    """Safely convert value to string."""
    try:
        if value is None:
            return default
        return str(value)
    except:
        return default

def create_optimised_cutlist_tab(ws, boards, core_db, laminate_db):
    """Create the main Optimised Cutlist tab matching user's desired format."""
    try:
        from openpyxl.styles import Font, PatternFill
        
        # Headers matching the new format with all fields for Optimised Cutlist (Column A-U alignment)
        headers = [
            'CLIENT NAME',           # A
            'ORDER ID / UNIQUE CODE', # B
            'SL NO.',               # C
            'ROOM TYPE',            # D
            'SUB CATEGORY',         # E
            'TYPE',                 # F
            'PANEL NAME',           # G
            'FULL NAME DESCRIPTION', # H
            'QTY',                  # I
            'GROOVE',               # J
            'CUT LENGTH',           # K
            'CUT WIDTH',            # L
            'FINISHED THICKNESS',    # M
            'MATERIAL TYPE',        # N
            'EB1',                  # O
            'EB2',                  # P
            'EB3',                  # Q
            'EB4',                  # R
            'GRAINS',               # S
            'REMARKS',              # T
            'Board ID',             # U
            'Material Length',      # V
            'Material Width',       # W
            'Length',              # X
            'Width',               # Y
            'Position (x)',        # Z
            'Position (y)',        # AA
            'Rotated',             # AB
            'Original Material',   # AC
            'Upgraded Material',   # AD
            'Grain Direction',     # AE
            'Material Upgraded'    # AF
        ]
        
        # Set headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=safe_str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        row = 2
        for board in boards:
            if not hasattr(board, 'parts_on_board') or not board.parts_on_board:
                continue
                
            # Get board material info
            board_material = safe_str(getattr(board, 'material_type', 'Unknown'))
            board_length = safe_float(getattr(board, 'length', 0))
            board_width = safe_float(getattr(board, 'width', 0))
            board_id = safe_str(getattr(board, 'id', f'Board_{row-1}'))
            
            for part in board.parts_on_board:
                # Get original CSV data that was stored in the part object
                def get_original_field(field_name, default=''):
                    """Get original CSV field data from part object"""
                    # Try original_data first (most reliable)
                    if hasattr(part, 'original_data') and field_name in part.original_data:
                        val = part.original_data[field_name]
                        return val if val and str(val).lower() != 'nan' else default
                    # Try direct attribute as fallback
                    attr_name = field_name.lower().replace(' ', '_').replace('/', '_').replace('.', '_')
                    if hasattr(part, attr_name):
                        return getattr(part, attr_name, default)
                    return default
                
                # New format fields (columns A-T, 1-20) - exact mapping to CSV headers
                ws.cell(row=row, column=1, value=safe_str(getattr(part, 'client_name', '')))                    # A: CLIENT NAME
                ws.cell(row=row, column=2, value=safe_str(getattr(part, 'id', 'Unknown')))                      # B: ORDER ID / UNIQUE CODE
                ws.cell(row=row, column=3, value=safe_str(get_original_field('SL NO.', str(getattr(part, 'original_part_index', '') or ''))))  # C: SL NO.
                ws.cell(row=row, column=4, value=safe_str(getattr(part, 'room_type', '')))                      # D: ROOM TYPE
                ws.cell(row=row, column=5, value=safe_str(get_original_field('SUB CATEGORY', getattr(part, 'sub_category', ''))))  # E: SUB CATEGORY
                ws.cell(row=row, column=6, value=safe_str(get_original_field('TYPE', 'Panel')))                 # F: TYPE
                ws.cell(row=row, column=7, value=safe_str(get_original_field('PANEL NAME', getattr(part, 'panel_name', ''))))  # G: PANEL NAME
                ws.cell(row=row, column=8, value=safe_str(get_original_field('FULL NAME DESCRIPTION', getattr(part, 'full_description', ''))))  # H: FULL NAME DESCRIPTION
                ws.cell(row=row, column=9, value=safe_int(get_original_field('QTY', getattr(part, 'quantity', 1))))  # I: QTY
                ws.cell(row=row, column=10, value=safe_str(get_original_field('GROOVE', '')))                   # J: GROOVE
                ws.cell(row=row, column=11, value=safe_float(getattr(part, 'requested_length', 0)))             # K: CUT LENGTH
                ws.cell(row=row, column=12, value=safe_float(getattr(part, 'requested_width', 0)))              # L: CUT WIDTH
                ws.cell(row=row, column=13, value=safe_str(get_original_field('FINISHED THICKNESS', '18')))     # M: FINISHED THICKNESS
                ws.cell(row=row, column=14, value=safe_str(getattr(part, 'material_details', '')))              # N: MATERIAL TYPE
                ws.cell(row=row, column=15, value=safe_str(get_original_field('EB1', '')))                      # O: EB1
                ws.cell(row=row, column=16, value=safe_str(get_original_field('EB2', '')))                      # P: EB2
                ws.cell(row=row, column=17, value=safe_str(get_original_field('EB3', '')))                      # Q: EB3
                ws.cell(row=row, column=18, value=safe_str(get_original_field('EB4', '')))                      # R: EB4
                ws.cell(row=row, column=19, value=safe_int(getattr(part, 'grains', 0)))                         # S: GRAINS
                ws.cell(row=row, column=20, value=safe_str(get_original_field('REMARKS', '')))                  # T: REMARKS
                
                # Optimization fields (columns U onward, 21+)
                ws.cell(row=row, column=21, value=board_id)  # U: Board ID
                
                # Material Length/Width - Original requested dimensions (before rotation)
                original_length = safe_float(getattr(part, 'requested_length', 0))
                original_width = safe_float(getattr(part, 'requested_width', 0))
                ws.cell(row=row, column=22, value=original_length)  # V: Material Length
                ws.cell(row=row, column=23, value=original_width)   # W: Material Width
                
                # Length/Width - Actual placed dimensions (after rotation if applicable)
                placed_length = original_length
                placed_width = original_width
                is_rotated = getattr(part, 'rotated', False)
                if is_rotated:
                    # Swap dimensions if rotated
                    placed_length = original_width
                    placed_width = original_length
                
                ws.cell(row=row, column=24, value=placed_length)   # X: Length
                ws.cell(row=row, column=25, value=placed_width)    # Y: Width
                
                # Position coordinates
                x_pos = safe_float(getattr(part, 'x_pos', getattr(part, 'x', 0)))
                y_pos = safe_float(getattr(part, 'y_pos', getattr(part, 'y', 0)))
                ws.cell(row=row, column=26, value=x_pos)  # Z: Position (x)
                ws.cell(row=row, column=27, value=y_pos)  # AA: Position (y)
                
                # Rotation status
                ws.cell(row=row, column=28, value='Yes' if is_rotated else 'No')  # AB: Rotated
                
                # Original Material - Get from part's original material details
                original_material = 'Unknown'
                if hasattr(part, 'material_details') and part.material_details:
                    original_material = safe_str(getattr(part.material_details, 'full_material_string', 'Unknown'))
                
                # Get board material from board.material_details.full_material_string
                board_material_final = 'Unknown'
                if hasattr(board, 'material_details') and board.material_details:
                    if hasattr(board.material_details, 'full_material_string'):
                        board_material_final = safe_str(board.material_details.full_material_string)
                    else:
                        board_material_final = safe_str(board.material_details)
                
                # Check if material was actually upgraded by comparing materials
                material_was_upgraded = False
                upgraded_material = ''
                
                # Debug logging - check board attributes
                board_attrs = [attr for attr in dir(board) if not attr.startswith('_')]
                logging.info(f"Board attributes: {board_attrs}")
                logging.info(f"Part {getattr(part, 'id', 'Unknown')}: original='{original_material}', board='{board_material_final}'")
                
                # Method 1: Compare full material strings
                if original_material != 'Unknown' and board_material_final != 'Unknown':
                    if original_material != board_material_final:
                        material_was_upgraded = True
                        upgraded_material = board_material_final
                        logging.info(f"UPGRADE DETECTED: {original_material} -> {board_material_final}")
                
                # Method 2: If no upgrade detected, check if part has upgrade info
                if not material_was_upgraded and hasattr(part, 'upgraded_material') and part.upgraded_material:
                    material_was_upgraded = True
                    upgraded_material = safe_str(part.upgraded_material)
                
                # Method 3: Enhanced core material extraction and comparison
                if not material_was_upgraded:
                    # Extract core from material strings using comprehensive patterns
                    original_core = ''
                    board_core = ''
                    
                    # Enhanced core extraction for original material
                    if '_' in original_material:
                        parts = original_material.split('_')
                        for part_str in parts:
                            # Look for standard core patterns: 18MDF, 18HDHMR, 16BWP, 6HDHMR, etc.
                            if any(core in part_str.upper() for core in ['MDF', 'MR', 'BWR', 'HDHMR', 'BWP']):
                                original_core = part_str
                                break
                    
                    # Enhanced core extraction for board material
                    if '_' in board_material_final:
                        parts = board_material_final.split('_')
                        for part_str in parts:
                            if any(core in part_str.upper() for core in ['MDF', 'MR', 'BWR', 'HDHMR', 'BWP']):
                                board_core = part_str
                                break
                    
                    logging.info(f"Enhanced core comparison: original='{original_core}', board='{board_core}'")
                    
                    # Check for actual material upgrade (different cores)
                    if original_core and board_core and original_core != board_core:
                        # Comprehensive upgrade patterns including BWP and various thicknesses
                        upgrade_patterns = [
                            ('18MDF', '18HDHMR'), ('18MR', '18HDHMR'), ('18BWR', '18HDHMR'),
                            ('16BWP', '18HDHMR'), ('16BWP', '18BWR'), ('16BWP', '18MR'),
                            ('18MR', '18BWR'), ('18MDF', '18BWR'), ('18MDF', '18MR'),
                            ('8MDF', '8HDHMR'), ('8MR', '8HDHMR'), ('6HDHMR', '8HDHMR'),
                            ('6HDHMR', '18HDHMR'), ('8HDHMR', '18HDHMR')
                        ]
                        for lower, higher in upgrade_patterns:
                            if lower.upper() in original_core.upper() and higher.upper() in board_core.upper():
                                material_was_upgraded = True
                                upgraded_material = board_material_final
                                logging.info(f"UPGRADE PATTERN DETECTED: {lower} -> {higher}")
                                break
                
                # Material tracking columns (AC-AF, 29-32)
                ws.cell(row=row, column=29, value=original_material)  # AC: Original Material
                
                # Set Upgraded Material and Material Upgraded columns
                if material_was_upgraded:
                    ws.cell(row=row, column=30, value=upgraded_material if upgraded_material else board_material_final)  # AD: Upgraded Material
                    ws.cell(row=row, column=32, value='Yes')  # AF: Material Upgraded
                else:
                    ws.cell(row=row, column=30, value=original_material)  # AD: Show original material when no upgrade
                    ws.cell(row=row, column=32, value='No')  # AF: Material Upgraded
                
                # Grain Direction - Check if part is grain sensitive
                grain_sensitive = safe_int(getattr(part, 'grains', 0))
                if grain_sensitive == 1:
                    ws.cell(row=row, column=31, value='Grain Sensitive')  # AE: Grain Direction
                else:
                    ws.cell(row=row, column=31, value='Not Grain Sensitive')  # AE: Grain Direction
                
                row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[column_letter].width = 15
            
    except Exception as e:
        logging.error(f"Error creating optimised cutlist tab: {e}")
        ws.cell(row=1, column=1, value=f"Error generating cutlist: {safe_str(e)}")

def create_core_material_report_tab(ws, boards, core_db, laminate_db):
    """Create Core Material Report tab matching user's format."""
    try:
        from openpyxl.styles import Font, PatternFill
        
        # Headers matching your exact format
        headers = [
            'Core Material', 'Board Count', 'Standard Area (sqft)', 'Utilized Area (sqft)', 
            'Wastage Area (sqft)', 'Utilization %', 'Wastage %', 'Unit Price (₹/sqft)', 'Total Cost (₹)'
        ]
        
        # Set headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=safe_str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Collect data by core material
        core_data = {}
        
        for board in boards:
            # Get core material from board's material details with enhanced extraction
            core_material = 'Unknown'
            if hasattr(board, 'material_details') and board.material_details:
                # Method 1: Direct core material attribute
                if hasattr(board.material_details, 'core_material'):
                    core_material = board.material_details.core_material
                elif hasattr(board.material_details, 'core_name'):
                    core_material = board.material_details.core_name
                # Method 2: Extract from full material string
                elif hasattr(board.material_details, 'full_material_string'):
                    material_string = board.material_details.full_material_string
                    if '_' in material_string:
                        parts = material_string.split('_')
                        for part in parts:
                            if any(core in part.upper() for core in ['MDF', 'MR', 'BWR', 'HDHMR', 'BWP']):
                                core_material = part
                                break
            
            if core_material not in core_data:
                core_data[core_material] = {
                    'board_count': 0,
                    'standard_area': 0,
                    'utilized_area': 0,
                    'total_cost': 0
                }
            
            # Calculate areas using correct board dimensions - fix conversion factor
            board_length = safe_float(getattr(board, 'total_length', 0))
            board_width = safe_float(getattr(board, 'total_width', 0))
            standard_area = (board_length * board_width) / 92903.04  # Convert mm² to sqft (1 sqft = 92903.04 mm²)
            
            utilized_area = 0
            if hasattr(board, 'parts_on_board') and board.parts_on_board:
                for part in board.parts_on_board:
                    part_length = safe_float(getattr(part, 'requested_length', 0))
                    part_width = safe_float(getattr(part, 'requested_width', 0))
                    utilized_area += (part_length * part_width) / 92903.04  # Convert mm² to sqft
            
            # Get pricing - fix key name and matching logic
            unit_price = 0
            for core_name, core_info in core_db.items():
                # Try exact match first, then partial match
                if core_name == core_material or (core_name in core_material and core_name != 'Unknown'):
                    if isinstance(core_info, dict):
                        # Fix key name: database uses 'price_per_sqm' not 'Price per SqM'
                        unit_price = safe_float(core_info.get('price_per_sqm', 0)) / 10.764  # Convert ₹/m² to ₹/sqft
                    else:
                        unit_price = safe_float(core_info) / 10.764
                    break
            
            core_data[core_material]['board_count'] += 1
            core_data[core_material]['standard_area'] += standard_area
            core_data[core_material]['utilized_area'] += utilized_area
            core_data[core_material]['total_cost'] += standard_area * unit_price
        
        # Write data rows
        row = 2
        for core_material, data in core_data.items():
            ws.cell(row=row, column=1, value=core_material)
            ws.cell(row=row, column=2, value=data['board_count'])
            ws.cell(row=row, column=3, value=round(data['standard_area'], 2))
            ws.cell(row=row, column=4, value=round(data['utilized_area'], 2))
            
            wastage_area = data['standard_area'] - data['utilized_area']
            ws.cell(row=row, column=5, value=round(wastage_area, 2))
            
            utilization_pct = (data['utilized_area'] / data['standard_area'] * 100) if data['standard_area'] > 0 else 0
            wastage_pct = 100 - utilization_pct
            
            ws.cell(row=row, column=6, value=f"{utilization_pct:.1f}%")
            ws.cell(row=row, column=7, value=f"{wastage_pct:.1f}%")
            
            # Get unit price for display - fix key name
            unit_price_display = 0
            for core_name, core_info in core_db.items():
                if core_name == core_material or (core_name in core_material and core_name != 'Unknown'):
                    if isinstance(core_info, dict):
                        unit_price_display = safe_float(core_info.get('price_per_sqm', 0)) / 10.764  # Convert ₹/m² to ₹/sqft
                    else:
                        unit_price_display = safe_float(core_info) / 10.764
                    break
            
            ws.cell(row=row, column=8, value=f"₹{unit_price_display:.2f}")
            ws.cell(row=row, column=9, value=f"₹{data['total_cost']:.2f}")
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
            
    except Exception as e:
        logging.error(f"Error creating core material report: {e}")
        ws.cell(row=1, column=1, value=f"Error generating core material report: {safe_str(e)}")

def create_laminate_report_tab(ws, boards, core_db, laminate_db):
    """Create Laminate Report tab counting top and bottom laminates separately."""
    try:
        from openpyxl.styles import Font, PatternFill
        
        # Headers
        headers = [
            'Laminate Type', 'Laminate Count', 'Standard Area (sqft)', 'Utilized Area (sqft)', 
            'Wastage Area (sqft)', 'Utilization %', 'Wastage %', 'Unit Price (₹/sqft)', 'Total Cost (₹)'
        ]
        
        # Set headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=safe_str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Collect data by laminate type - count top and bottom separately
        laminate_data = {}
        
        for board in boards:
            # Calculate board areas first
            board_length = safe_float(getattr(board, 'total_length', 0))
            board_width = safe_float(getattr(board, 'total_width', 0))
            standard_area = (board_length * board_width) / 92903.04  # Convert mm² to sqft
            
            utilized_area = 0
            if hasattr(board, 'parts_on_board') and board.parts_on_board:
                for part in board.parts_on_board:
                    part_length = safe_float(getattr(part, 'requested_length', 0))
                    part_width = safe_float(getattr(part, 'requested_width', 0))
                    utilized_area += (part_length * part_width) / 92903.04  # Convert mm² to sqft
            
            # Process both top and bottom laminates separately
            if hasattr(board, 'material_details') and board.material_details:
                # Top laminate
                top_laminate = 'Unknown'
                if hasattr(board.material_details, 'top_laminate_name'):
                    top_laminate = board.material_details.top_laminate_name
                elif hasattr(board.material_details, 'laminate_name'):
                    top_laminate = board.material_details.laminate_name
                
                # Bottom laminate
                bottom_laminate = 'Unknown'
                if hasattr(board.material_details, 'bottom_laminate_name'):
                    bottom_laminate = board.material_details.bottom_laminate_name
                elif hasattr(board.material_details, 'laminate_name'):
                    bottom_laminate = board.material_details.laminate_name
                
                # Count each laminate separately
                for laminate_name in [top_laminate, bottom_laminate]:
                    if laminate_name and laminate_name != 'Unknown':
                        if laminate_name not in laminate_data:
                            laminate_data[laminate_name] = {
                                'board_count': 0,
                                'standard_area': 0,
                                'utilized_area': 0,
                                'total_cost': 0
                            }
                        
                        # Get pricing - find best match in laminate_db
                        unit_price = 0
                        for lam_name, lam_price in laminate_db.items():
                            if lam_name == laminate_name or (lam_name in laminate_name and lam_name != 'Unknown'):
                                unit_price = safe_float(lam_price) / 10.764  # Convert ₹/m² to ₹/sqft
                                break
                        
                        laminate_data[laminate_name]['board_count'] += 1
                        laminate_data[laminate_name]['standard_area'] += standard_area
                        laminate_data[laminate_name]['utilized_area'] += utilized_area
                        laminate_data[laminate_name]['total_cost'] += standard_area * unit_price
        
        # Write data rows
        row = 2
        for laminate_type, data in laminate_data.items():
            ws.cell(row=row, column=1, value=laminate_type)
            ws.cell(row=row, column=2, value=data['board_count'])
            ws.cell(row=row, column=3, value=round(data['standard_area'], 2))
            ws.cell(row=row, column=4, value=round(data['utilized_area'], 2))
            
            wastage_area = data['standard_area'] - data['utilized_area']
            ws.cell(row=row, column=5, value=round(wastage_area, 2))
            
            utilization_pct = (data['utilized_area'] / data['standard_area'] * 100) if data['standard_area'] > 0 else 0
            wastage_pct = 100 - utilization_pct
            
            ws.cell(row=row, column=6, value=f"{utilization_pct:.1f}%")
            ws.cell(row=row, column=7, value=f"{wastage_pct:.1f}%")
            
            # Get unit price for display
            unit_price_display = 0
            for lam_name, lam_price in laminate_db.items():
                if lam_name == laminate_type or (lam_name in laminate_type and lam_name != 'Unknown'):
                    unit_price_display = safe_float(lam_price) / 10.764  # Convert ₹/m² to ₹/sqft
                    break
            
            ws.cell(row=row, column=8, value=f"₹{unit_price_display:.2f}")
            ws.cell(row=row, column=9, value=f"₹{data['total_cost']:.2f}")
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
            
    except Exception as e:
        logging.error(f"Error creating laminate report: {e}")
        ws.cell(row=1, column=1, value=f"Error generating laminate report: {safe_str(e)}")

def create_material_upgrade_report_tab(ws, boards, upgrade_summary):
    """Create Material Upgrade Report tab with comprehensive upgrade detection."""
    try:
        from openpyxl.styles import Font, PatternFill
        
        # Headers
        headers = ['Original Material', 'Upgraded Material', 'Parts Count', 'Upgrade Type']
        
        # Set headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=safe_str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Collect upgrade data with proper board material access
        upgrade_data = {}
        
        for board in boards:
            if not hasattr(board, 'parts_on_board') or not board.parts_on_board:
                continue
                
            # Get board material from correct attribute
            board_material = 'Unknown'
            if hasattr(board, 'material_details') and board.material_details:
                if hasattr(board.material_details, 'full_material_string'):
                    board_material = safe_str(board.material_details.full_material_string)
                else:
                    board_material = safe_str(board.material_details)
            
            for part in board.parts_on_board:
                # Get original material from part
                original_material = 'Unknown'
                if hasattr(part, 'material_details') and part.material_details:
                    original_material = safe_str(getattr(part.material_details, 'full_material_string', 'Unknown'))
                
                # Determine if upgrade occurred and what type
                upgrade_type = 'No Upgrade'
                if original_material != 'Unknown' and board_material != 'Unknown':
                    if original_material != board_material:
                        # Extract cores for upgrade type determination
                        original_core = ''
                        board_core = ''
                        
                        if '_' in original_material:
                            parts = original_material.split('_')
                            for part_str in parts:
                                if any(core in part_str.upper() for core in ['MDF', 'MR', 'BWR', 'HDHMR', 'BWP']):
                                    original_core = part_str
                                    break
                        
                        if '_' in board_material:
                            parts = board_material.split('_')
                            for part_str in parts:
                                if any(core in part_str.upper() for core in ['MDF', 'MR', 'BWR', 'HDHMR', 'BWP']):
                                    board_core = part_str
                                    break
                        
                        # Determine upgrade type
                        if original_core and board_core:
                            if 'MDF' in original_core.upper() and 'HDHMR' in board_core.upper():
                                upgrade_type = 'MDF to HDHMR'
                            elif 'MR' in original_core.upper() and 'HDHMR' in board_core.upper():
                                upgrade_type = 'MR to HDHMR'
                            elif 'BWP' in original_core.upper() and ('HDHMR' in board_core.upper() or 'BWR' in board_core.upper()):
                                upgrade_type = 'BWP to Higher Grade'
                            else:
                                upgrade_type = 'Core Upgrade'
                        else:
                            upgrade_type = 'Material Upgrade'
                
                # Create upgrade key
                if upgrade_type != 'No Upgrade':
                    upgrade_key = (original_material, board_material, upgrade_type)
                    
                    if upgrade_key not in upgrade_data:
                        upgrade_data[upgrade_key] = 0
                    upgrade_data[upgrade_key] += 1
        
        # Write data rows
        row = 2
        for (original, upgraded, upgrade_type), count in upgrade_data.items():
            ws.cell(row=row, column=1, value=original)
            ws.cell(row=row, column=2, value=upgraded)
            ws.cell(row=row, column=3, value=count)
            ws.cell(row=row, column=4, value=upgrade_type)
            row += 1
        
        # If no upgrades found, show message
        if row == 2:
            ws.cell(row=2, column=1, value="No material upgrades detected")
            ws.cell(row=2, column=2, value="All parts used original materials")
            ws.cell(row=2, column=3, value="0")
            ws.cell(row=2, column=4, value="No Upgrade")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 25
            
    except Exception as e:
        logging.error(f"Error creating material upgrade report: {e}")
        ws.cell(row=1, column=1, value=f"Error generating material upgrade report: {safe_str(e)}")

def create_edge_band_summary_tab(ws, boards):
    """Create Edge Band Summary tab for Excel report."""
    try:
        from openpyxl.styles import Font, PatternFill
        
        # Headers for edge band report
        headers = [
            'Edge Band Name', 'Panel Count', 'Total Length (mm)', 'Total Length (m)'
        ]
        
        # Set headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=safe_str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Collect edge band data by top laminate
        edge_band_data = {}
        
        for board in boards:
            if not hasattr(board, 'parts_on_board') or not board.parts_on_board:
                continue
                
            for part in board.parts_on_board:
                # Get top laminate from part material details
                top_laminate = 'Unknown'
                
                if hasattr(part, 'material_details') and part.material_details:
                    if hasattr(part.material_details, 'top_laminate_name'):
                        top_laminate = safe_str(part.material_details.top_laminate_name)
                    elif hasattr(part.material_details, 'laminate_material'):
                        top_laminate = safe_str(part.material_details.laminate_material)
                    elif hasattr(part.material_details, 'top_laminate'):
                        top_laminate = safe_str(part.material_details.top_laminate)
                    elif hasattr(part.material_details, 'full_material_string'):
                        # Extract laminate from material string format: laminate_core_laminate
                        material_string = safe_str(part.material_details.full_material_string)
                        if '_' in material_string:
                            parts = material_string.split('_')
                            if len(parts) >= 3:
                                top_laminate = parts[0]  # First part is top laminate
                
                # Calculate perimeter of the panel (2 * length + 2 * width)
                part_length = safe_float(getattr(part, 'requested_length', 0))
                part_width = safe_float(getattr(part, 'requested_width', 0))
                perimeter_mm = 2 * (part_length + part_width)
                
                # Group by top laminate
                if top_laminate not in edge_band_data:
                    edge_band_data[top_laminate] = {
                        'panel_count': 0,
                        'total_perimeter_mm': 0
                    }
                
                edge_band_data[top_laminate]['panel_count'] += 1
                edge_band_data[top_laminate]['total_perimeter_mm'] += perimeter_mm
        
        # Write data rows
        row = 2
        for laminate, data in sorted(edge_band_data.items()):
            # Convert mm to meters for better readability
            total_length_m = data['total_perimeter_mm'] / 1000
            
            ws.cell(row=row, column=1, value=laminate)
            ws.cell(row=row, column=2, value=data['panel_count'])
            ws.cell(row=row, column=3, value=data['total_perimeter_mm'])
            ws.cell(row=row, column=4, value=round(total_length_m, 2))
            row += 1
        
        # If no data found, show message
        if row == 2:
            ws.cell(row=2, column=1, value="No edge band data available")
            ws.cell(row=2, column=2, value="0")
            ws.cell(row=2, column=3, value="0")
            ws.cell(row=2, column=4, value="0.00")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
            
    except Exception as e:
        logging.error(f"Error creating edge band summary tab: {e}")
        ws.cell(row=1, column=1, value=f"Error generating edge band summary: {safe_str(e)}")

def create_excel_report_robust(boards, unplaced_parts, upgrade_summary, initial_cost, final_cost, 
                              order_name, core_db, laminate_db):
    """Create comprehensive Excel report with robust error handling."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        
        # Remove default sheet safely - fix openpyxl compatibility
        try:
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
        except Exception:
            # If removal fails, just continue with existing sheets
            pass
        
        # Tab 1: Optimised Cutlist (main detailed view)
        ws_cutlist = wb.create_sheet("Optimised Cutlist", 0)
        create_optimised_cutlist_tab(ws_cutlist, boards, core_db, laminate_db)
        
        # Tab 2: Summary
        ws_summary = wb.create_sheet("Summary", 1)
        create_summary_tab_safe(ws_summary, boards, unplaced_parts, initial_cost, final_cost, order_name)
        
        # Tab 3: Board Details
        ws_boards = wb.create_sheet("Board Details", 2)
        create_board_details_tab_safe(ws_boards, boards, core_db, laminate_db)
        
        # Tab 4: Parts List
        ws_parts = wb.create_sheet("Parts List", 3)
        create_parts_list_tab_safe(ws_parts, boards, unplaced_parts)
        
        # Tab 5: Core Material Report
        ws_core_materials = wb.create_sheet("Core Material Report", 4)
        create_core_material_report_tab(ws_core_materials, boards, core_db, laminate_db)
        
        # Tab 6: Laminate Report
        ws_laminate_report = wb.create_sheet("Laminate Report", 5)
        create_laminate_report_tab(ws_laminate_report, boards, core_db, laminate_db)
        
        # Tab 7: Edge Band Summary
        ws_edge_band = wb.create_sheet("Edge Band Summary", 6)
        create_edge_band_summary_tab(ws_edge_band, boards)
        
        # Tab 8: Material Upgrade Report
        ws_upgrade_report = wb.create_sheet("Material Upgrade Report", 7)
        create_material_upgrade_report_tab(ws_upgrade_report, boards, upgrade_summary)
        
        # Tab 9: Material Summary (original)
        ws_materials = wb.create_sheet("Material Summary", 8)
        create_material_summary_tab_safe(ws_materials, boards, core_db, laminate_db)
        
        # Tab 10: Cost Analysis
        ws_cost = wb.create_sheet("Cost Analysis", 9)
        create_cost_analysis_tab_safe(ws_cost, boards, core_db, laminate_db, initial_cost, final_cost)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
        
    except Exception as e:
        logging.error(f"Excel generation failed: {e}")
        raise

def create_summary_tab_safe(ws, boards, unplaced_parts, initial_cost, final_cost, order_name):
    """Create summary tab with safe data handling."""
    try:
        # Title
        ws['A1'] = safe_str(f"OptiWise Optimization Summary - {order_name}")
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Calculate metrics safely
        total_parts = safe_int(len(unplaced_parts))
        placed_parts = 0
        total_utilization = 0.0
        
        for board in boards:
            if hasattr(board, 'parts_on_board'):
                placed_parts += safe_int(len(board.parts_on_board))
                total_parts += safe_int(len(board.parts_on_board))
            if hasattr(board, 'get_utilization_percentage'):
                total_utilization += safe_float(board.get_utilization_percentage())
        
        avg_utilization = total_utilization / max(len(boards), 1) if boards else 0
        
        initial_cost_safe = safe_float(initial_cost)
        final_cost_safe = safe_float(final_cost)
        savings = initial_cost_safe - final_cost_safe
        savings_pct = (savings / initial_cost_safe * 100) if initial_cost_safe > 0 else 0
        
        # Write metrics
        row = 3
        metrics = [
            ("Total Boards Used", safe_int(len(boards))),
            ("Total Parts", total_parts),
            ("Parts Placed", placed_parts),
            ("Parts Unplaced", safe_int(len(unplaced_parts))),
            ("Average Utilization (%)", safe_float(avg_utilization, 2)),
            ("Initial Cost", initial_cost_safe),
            ("Final Cost", final_cost_safe),
            ("Cost Savings", savings),
            ("Savings Percentage", savings_pct)
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = safe_str(metric)
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
    except Exception as e:
        logging.error(f"Summary tab creation failed: {e}")
        ws['A1'] = "Error creating summary"

def create_board_details_tab_safe(ws, boards, core_db, laminate_db):
    """Create board details tab with safe data handling."""
    try:
        # Headers
        headers = ['Board ID', 'Material', 'Size (mm)', 'Utilization %', 'Parts Count', 'Cost', 'Waste Area']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=safe_str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Board data
        for row, board in enumerate(boards, 2):
            try:
                board_id = safe_str(getattr(board, 'id', f'Board_{row-1}'))
                material = safe_str(getattr(board, 'material_details', 'Unknown'))
                
                length = safe_float(getattr(board, 'total_length', 0))
                width = safe_float(getattr(board, 'total_width', 0))
                size_str = f"{length:.0f}x{width:.0f}"
                
                utilization = safe_float(board.get_utilization_percentage() if hasattr(board, 'get_utilization_percentage') else 0)
                parts_count = safe_int(len(getattr(board, 'parts_on_board', [])))
                
                # Calculate cost safely
                board_cost = calculate_board_cost_safe(board, core_db, laminate_db)
                
                # Calculate waste
                total_area = length * width
                utilized_area = total_area * utilization / 100
                waste_area = (total_area - utilized_area) / 1_000_000  # Convert to sqm
                
                ws.cell(row=row, column=1, value=board_id)
                ws.cell(row=row, column=2, value=material)
                ws.cell(row=row, column=3, value=size_str)
                ws.cell(row=row, column=4, value=utilization)
                ws.cell(row=row, column=5, value=parts_count)
                ws.cell(row=row, column=6, value=board_cost)
                ws.cell(row=row, column=7, value=safe_float(waste_area))
                
            except Exception as e:
                logging.warning(f"Error processing board {row}: {e}")
                continue
                
    except Exception as e:
        logging.error(f"Board details tab creation failed: {e}")
        ws['A1'] = "Error creating board details"

def create_parts_list_tab_safe(ws, boards, unplaced_parts):
    """Create parts list tab with safe data handling."""
    try:
        # Headers
        headers = ['Part ID', 'Size (mm)', 'Material', 'Board ID', 'Position', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=safe_str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row = 2
        # Placed parts
        for board in boards:
            try:
                board_id = safe_str(getattr(board, 'id', 'Unknown'))
                parts_on_board = getattr(board, 'parts_on_board', [])
                
                for part in parts_on_board:
                    try:
                        part_id = safe_str(getattr(part, 'id', 'Unknown'))
                        length = safe_float(getattr(part, 'requested_length', 0))
                        width = safe_float(getattr(part, 'requested_width', 0))
                        size_str = f"{length:.0f}x{width:.0f}"
                        material = safe_str(getattr(part, 'material_details', 'Unknown'))
                        
                        x_pos = safe_float(getattr(part, 'x_pos', 0))
                        y_pos = safe_float(getattr(part, 'y_pos', 0))
                        position = f"({x_pos:.0f},{y_pos:.0f})"
                        
                        ws.cell(row=row, column=1, value=part_id)
                        ws.cell(row=row, column=2, value=size_str)
                        ws.cell(row=row, column=3, value=material)
                        ws.cell(row=row, column=4, value=board_id)
                        ws.cell(row=row, column=5, value=position)
                        ws.cell(row=row, column=6, value="Placed")
                        row += 1
                        
                    except Exception as e:
                        logging.warning(f"Error processing part: {e}")
                        continue
                        
            except Exception as e:
                logging.warning(f"Error processing board parts: {e}")
                continue
        
        # Unplaced parts
        for part in unplaced_parts:
            try:
                part_id = safe_str(getattr(part, 'id', 'Unknown'))
                length = safe_float(getattr(part, 'requested_length', 0))
                width = safe_float(getattr(part, 'requested_width', 0))
                size_str = f"{length:.0f}x{width:.0f}"
                material = safe_str(getattr(part, 'material_details', 'Unknown'))
                
                ws.cell(row=row, column=1, value=part_id)
                ws.cell(row=row, column=2, value=size_str)
                ws.cell(row=row, column=3, value=material)
                ws.cell(row=row, column=4, value="N/A")
                ws.cell(row=row, column=5, value="N/A")
                ws.cell(row=row, column=6, value="Unplaced")
                row += 1
                
            except Exception as e:
                logging.warning(f"Error processing unplaced part: {e}")
                continue
                
    except Exception as e:
        logging.error(f"Parts list tab creation failed: {e}")
        ws['A1'] = "Error creating parts list"

def create_material_summary_tab_safe(ws, boards, core_db, laminate_db):
    """Create material summary tab with safe data handling."""
    try:
        # Headers
        headers = ['Material', 'Boards Used', 'Total Area (sqm)', 'Utilized Area (sqm)', 'Waste Area (sqm)', 'Utilization %', 'Total Cost']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=safe_str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Group by material
        material_summary = {}
        
        for board in boards:
            try:
                material_key = safe_str(getattr(board, 'material_details', 'Unknown'))
                
                if material_key not in material_summary:
                    material_summary[material_key] = {
                        'boards': 0, 'total_area': 0, 'utilized_area': 0, 'waste_area': 0, 'cost': 0
                    }
                
                length = safe_float(getattr(board, 'total_length', 0))
                width = safe_float(getattr(board, 'total_width', 0))
                board_area = length * width / 1_000_000  # Convert to sqm
                
                utilization = safe_float(board.get_utilization_percentage() if hasattr(board, 'get_utilization_percentage') else 0)
                utilized_area = board_area * utilization / 100
                waste_area = board_area - utilized_area
                
                board_cost = calculate_board_cost_safe(board, core_db, laminate_db)
                
                summary = material_summary[material_key]
                summary['boards'] += 1
                summary['total_area'] += board_area
                summary['utilized_area'] += utilized_area
                summary['waste_area'] += waste_area
                summary['cost'] += board_cost
                
            except Exception as e:
                logging.warning(f"Error processing board for material summary: {e}")
                continue
        
        # Write data
        row = 2
        for material, summary in material_summary.items():
            try:
                utilization = (summary['utilized_area'] / summary['total_area'] * 100) if summary['total_area'] > 0 else 0
                
                ws.cell(row=row, column=1, value=safe_str(material))
                ws.cell(row=row, column=2, value=safe_int(summary['boards']))
                ws.cell(row=row, column=3, value=safe_float(summary['total_area']))
                ws.cell(row=row, column=4, value=safe_float(summary['utilized_area']))
                ws.cell(row=row, column=5, value=safe_float(summary['waste_area']))
                ws.cell(row=row, column=6, value=safe_float(utilization))
                ws.cell(row=row, column=7, value=safe_float(summary['cost']))
                row += 1
                
            except Exception as e:
                logging.warning(f"Error writing material summary row: {e}")
                continue
                
    except Exception as e:
        logging.error(f"Material summary tab creation failed: {e}")
        ws['A1'] = "Error creating material summary"

def create_cost_analysis_tab_safe(ws, boards, core_db, laminate_db, initial_cost, final_cost):
    """Create cost analysis tab with safe data handling."""
    try:
        ws['A1'] = "Cost Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Cost summary
        initial_cost_safe = safe_float(initial_cost)
        final_cost_safe = safe_float(final_cost)
        savings = initial_cost_safe - final_cost_safe
        savings_pct = (savings / initial_cost_safe * 100) if initial_cost_safe > 0 else 0
        
        ws['A3'] = "Cost Summary"
        ws['A3'].font = Font(bold=True)
        ws['A4'] = "Initial Cost (Worst Case)"
        ws['B4'] = initial_cost_safe
        ws['A5'] = "Optimized Cost"
        ws['B5'] = final_cost_safe
        ws['A6'] = "Total Savings"
        ws['B6'] = savings
        ws['A7'] = "Savings Percentage"
        ws['B7'] = savings_pct
        
        # Board-wise cost breakdown
        ws['A9'] = "Board-wise Cost Breakdown"
        ws['A9'].font = Font(bold=True)
        
        headers = ['Board ID', 'Material', 'Core Cost', 'Laminate Cost', 'Total Cost']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col, value=safe_str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row = 11
        for board in boards:
            try:
                board_id = safe_str(getattr(board, 'id', f'Board_{row-10}'))
                material = safe_str(getattr(board, 'material_details', 'Unknown'))
                
                core_cost, laminate_cost = calculate_detailed_board_cost_safe(board, core_db, laminate_db)
                total_cost = core_cost + laminate_cost
                
                ws.cell(row=row, column=1, value=board_id)
                ws.cell(row=row, column=2, value=material)
                ws.cell(row=row, column=3, value=safe_float(core_cost))
                ws.cell(row=row, column=4, value=safe_float(laminate_cost))
                ws.cell(row=row, column=5, value=safe_float(total_cost))
                row += 1
                
            except Exception as e:
                logging.warning(f"Error processing board cost: {e}")
                continue
                
    except Exception as e:
        logging.error(f"Cost analysis tab creation failed: {e}")
        ws['A1'] = "Error creating cost analysis"

def calculate_board_cost_safe(board, core_db, laminate_db):
    """Calculate board cost with safe error handling."""
    try:
        core_cost, laminate_cost = calculate_detailed_board_cost_safe(board, core_db, laminate_db)
        return safe_float(core_cost + laminate_cost)
    except Exception as e:
        logging.warning(f"Error calculating board cost: {e}")
        return 0.0

def calculate_detailed_board_cost_safe(board, core_db, laminate_db):
    """Calculate detailed board cost breakdown with safe error handling."""
    try:
        length = safe_float(getattr(board, 'total_length', 0))
        width = safe_float(getattr(board, 'total_width', 0))
        board_area_sqm = length * width / 1_000_000
        
        core_cost = 0.0
        laminate_cost = 0.0
        
        # Core cost calculation
        if core_db and hasattr(board, 'material_details'):
            material_details = getattr(board, 'material_details', None)
            if material_details and hasattr(material_details, 'core_name'):
                core_name = getattr(material_details, 'core_name', '')
                if core_name in core_db:
                    core_info = core_db[core_name]
                    if isinstance(core_info, dict) and 'price_per_sqm' in core_info:
                        core_price = safe_float(core_info['price_per_sqm'])
                        core_cost = core_price * board_area_sqm
        
        # Laminate cost calculation - count both top and bottom separately
        if laminate_db and hasattr(board, 'material_details'):
            material_details = getattr(board, 'material_details', None)
            if material_details:
                # Extract laminates from full material string if direct attributes not available
                if hasattr(material_details, 'full_material_string'):
                    material_string = material_details.full_material_string
                    if '_' in material_string:
                        parts = material_string.split('_')
                        if len(parts) >= 3:  # Format: top_core_bottom
                            top_laminate = parts[0]
                            bottom_laminate = parts[2]
                            
                            # Top laminate cost
                            if top_laminate in laminate_db:
                                laminate_price = safe_float(laminate_db[top_laminate])
                                laminate_cost += laminate_price * board_area_sqm
                            
                            # Bottom laminate cost (always add, even if same as top)
                            if bottom_laminate in laminate_db:
                                laminate_price = safe_float(laminate_db[bottom_laminate])
                                laminate_cost += laminate_price * board_area_sqm
                
                # Fallback to direct attributes
                if laminate_cost == 0:
                    # Top laminate
                    if hasattr(material_details, 'top_laminate_name'):
                        top_laminate = getattr(material_details, 'top_laminate_name', '')
                        if top_laminate in laminate_db:
                            laminate_price = safe_float(laminate_db[top_laminate])
                            laminate_cost += laminate_price * board_area_sqm
                    
                    # Bottom laminate (always add separately)
                    if hasattr(material_details, 'bottom_laminate_name'):
                        bottom_laminate = getattr(material_details, 'bottom_laminate_name', '')
                        if bottom_laminate in laminate_db:
                            laminate_price = safe_float(laminate_db[bottom_laminate])
                            laminate_cost += laminate_price * board_area_sqm
        
        return safe_float(core_cost), safe_float(laminate_cost)
        
    except Exception as e:
        logging.warning(f"Error in detailed cost calculation: {e}")
        return 0.0, 0.0
